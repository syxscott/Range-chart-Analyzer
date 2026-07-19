"""Table configuration + CSV / TSV / JSON / XLSX export helpers.

The table configs are the single source of truth for both the GUI tables
and the export columns, so exported columns always match what is shown.

Two modes: range-chart (default) and columnar-section. The shape of the
result dict selects which table set is used.

XLSX support requires ``openpyxl``; if it's missing the GUI falls back
to a clear error message so the user knows which dep to install.
"""

from __future__ import annotations

import csv
import io
from typing import Any, Callable

# Each config: id, i18n title key, list of column i18n keys, and a row
# extractor producing cell values in column order.
def _looks_columnar(data: dict[str, Any] | None) -> bool:
    """Heuristic: if data.sections is present and its first item has an 'id'
    field (column-label) rather than a 'name' (measured-section), it's a
    columnar-section result. Mirrors js/table.js.
    """
    if not data:
        return False
    sects = data.get("sections")
    if not isinstance(sects, list):
        return False
    if not sects:
        return False
    first = sects[0]
    # "id" is the definitive columnar marker (it holds the column-label
    # value). Even if "name" is also present (LLM being verbose), the
    # presence of "id" is sufficient to classify as columnar. Range-chart
    # sections use "name" (measured-section name) and never have "id".
    return isinstance(first, dict) and ("id" in first)


def _looks_abundance(data: dict[str, Any] | None) -> bool:
    """Heuristic: an abundance-diagram result carries an ``abundances`` list
    (the per-(taxon, level) rows). Mirrors js/table.js.

    An empty list (`abundances: []`) is NOT enough to call this an
    abundance result — every freshly-extracted object schema includes an
    empty `abundances` placeholder before the model fills it in, so
    accepting empty lists misroutes pure range-chart results to the
    abundance renderer. Require at least one entry plus a non-empty
    `abundances` list to disambiguate.
    """
    if not data:
        return False
    ab = data.get("abundances")
    return isinstance(ab, list) and len(ab) > 0


def _range_chart_tables(data: dict[str, Any] | None) -> list[dict[str, Any]]:
    """Range-chart table configs. Multi-run results gain an agreement column
    on the species table — mirrors js/table.js.
    """
    multi = bool(data) and int(data.get("runs", 1) or 1) > 1
    species_cols = ["col.species", "col.section", "col.rangeBase", "col.rangeTop", "col.biozone"]
    species_data = ["species", "section", "range_base", "range_top", "biozone"]
    species_row_base = lambda r: [
        r.get("species", ""),
        r.get("section", ""),
        r.get("range_base", ""),
        r.get("range_top", ""),
        r.get("biozone", ""),
    ]

    return [
        {
            "id": "sections",
            "title_key": "sec.sections",
            "cols": ["col.name", "col.ageRange", "col.formations", "col.thickness", "col.coordinates"],
            "data_keys": ["name", "age_range", "formations", "formation_thickness_m", "coordinates"],
            "row": lambda s: [
                s.get("name", ""),
                s.get("age_range", ""),
                "; ".join(s.get("formations", []) or []),
                s.get("formation_thickness_m", ""),
                s.get("coordinates", ""),
            ],
        },
        {
            "id": "species_ranges",
            "title_key": "sec.species",
            "cols": species_cols + (["col.agreement"] if multi else []),
            "data_keys": species_data + (["agreement"] if multi else []),
            "row": (lambda r: species_row_base(r) + [r.get("agreement", "")]) if multi else species_row_base,
        },
        {
            "id": "biozones",
            "title_key": "sec.biozones",
            "cols": ["col.name", "col.section", "col.age", "col.thickness"],
            "data_keys": ["name", "section", "age", "thickness_m"],
            "row": lambda b: [b.get("name", ""), b.get("section", ""), b.get("age", ""), b.get("thickness_m", "")],
        },
        {
            "id": "other_fossils",
            "title_key": "sec.fossils",
            "cols": ["col.fossil"],
            "data_keys": ["fossil"],
            # Defensive: items may be plain strings (from rca_core
            # normalize_result) or dicts (older or third-party producers).
            "row": lambda f: [f.get("fossil", f.get("text", ""))] if isinstance(f, dict) else [f],
        },
    ]


def _columnar_section_tables(data: dict[str, Any] | None) -> list[dict[str, Any]]:
    """Columnar-section table configs. Multi-run results gain an agreement
    column on the sections table — mirrors js/table.js.
    """
    multi = bool(data) and int(data.get("runs", 1) or 1) > 1
    sec_cols = ["col.sectionId", "col.sectionGroup", "col.thickness", "col.coordinates"]
    sec_data = ["id", "group", "thickness_m", "coordinates_text"]
    cols_final = sec_cols + (["col.agreement"] if multi else [])
    data_final = sec_data + (["agreement"] if multi else [])
    row_base = lambda s: [
        s.get("id", ""),
        s.get("group", ""),
        s.get("thickness_m", ""),
        s.get("coordinates_text", ""),
    ]
    row = (lambda s: row_base(s) + [s.get("agreement", "")]) if multi else row_base

    return [
        {
            "id": "sections",
            "title_key": "sec.sections",
            "cols": cols_final,
            "data_keys": data_final,
            "row": row,
        },
        {
            "id": "fossil_legend",
            "title_key": "sec.fossils",
            "cols": ["col.fossilMarker", "col.fossilMeaning"],
            "data_keys": ["marker", "meaning"],
            "row": lambda it: [it.get("marker", ""), it.get("meaning", "")],
        },
        {
            "id": "lithology_legend",
            "title_key": "sec.columnarLithology",
            "cols": ["col.lithologyPattern", "col.lithologyMeaning"],
            "data_keys": ["pattern", "meaning"],
            "row": lambda it: [it.get("pattern", it.get("marker", "")), it.get("meaning", "")],
        },
        {
            "id": "cross_beds",
            "title_key": "sec.crossBeds",
            "cols": ["col.crossFrom", "col.crossFromBed", "col.crossTo", "col.crossToBed"],
            "data_keys": ["from_section", "from_bed_idx", "to_section", "to_bed_idx"],
            "row": lambda it: [
                it.get("from_section", ""),
                "" if it.get("from_bed_idx") is None else str(it.get("from_bed_idx")),
                it.get("to_section", ""),
                "" if it.get("to_bed_idx") is None else str(it.get("to_bed_idx")),
            ],
        },
    ]


def get_configs_for_result(data: dict[str, Any] | None) -> list[dict[str, Any]]:
    """Return the table config list appropriate for ``data``."""
    if _looks_abundance(data):
        return _abundance_diagram_tables(data)
    return _columnar_section_tables(data) if _looks_columnar(data) else _range_chart_tables(data)


def _abundance_diagram_tables(data: dict[str, Any] | None) -> list[dict[str, Any]]:
    """Abundance-diagram (pollen / percentage-diagram) table configs.

    Multi-run results gain an agreement column on the abundances table —
    mirrors js/table.js.
    """
    multi = bool(data) and int(data.get("runs", 1) or 1) > 1
    ab_cols = ["col.taxon", "col.site", "col.level", "col.depth", "col.abundance", "col.abundanceUnit"]
    ab_data = ["taxon", "site", "level", "depth", "abundance", "abundance_unit"]
    ab_row_base = lambda r: [
        r.get("taxon", ""),
        r.get("site", ""),
        r.get("level", ""),
        r.get("depth", ""),
        r.get("abundance", ""),
        r.get("abundance_unit", ""),
    ]

    return [
        {
            "id": "sites",
            "title_key": "sec.sites",
            "cols": ["col.name", "col.location", "col.ageRange", "col.depthUnit"],
            "data_keys": ["name", "location", "age_range", "depth_unit"],
            "row": lambda s: [
                s.get("name", ""),
                s.get("location", ""),
                s.get("age_range", ""),
                s.get("depth_unit", ""),
            ],
        },
        {
            "id": "abundances",
            "title_key": "sec.abundances",
            "cols": ab_cols + (["col.agreement"] if multi else []),
            "data_keys": ab_data + (["agreement"] if multi else []),
            "row": (lambda r: ab_row_base(r) + [r.get("agreement", "")]) if multi else ab_row_base,
        },
        {
            "id": "zones",
            "title_key": "sec.zones",
            "cols": ["col.name", "col.age", "col.levelRange"],
            "data_keys": ["name", "age", "level_range"],
            "row": lambda z: [z.get("name", ""), z.get("age", ""), z.get("level_range", "")],
        },
    ]


def get_config(table_id: str) -> dict[str, Any] | None:
    # Search through all presets — used by the fallback path in
    # build_table_export / apply_table_edits when the table isn't in the
    # data-shape-matched configs (e.g. editing a cross_beds row while the
    # result has no sections to detect columnar shape).
    for fn in (_range_chart_tables, _columnar_section_tables, _abundance_diagram_tables):
        for c in fn(None):
            if c["id"] == table_id:
                return c
    return None


def _find_cfg(table_id: str, data: dict[str, Any] | None) -> dict[str, Any] | None:
    """Resolve a table config: first the shape-matched configs (so the
    multi-run agreement column and columnar/range/abundance routing apply),
    then a cross-shape fallback so tables whose shape can't be inferred
    from the data (e.g. editing cross_beds with an empty sections list) are
    still found.
    """
    for c in get_configs_for_result(data):
        if c["id"] == table_id:
            return c
    return get_config(table_id)


# Legacy alias kept for callers that imported TABLE_CONFIGS directly.
TABLE_CONFIGS = _range_chart_tables(None)


def build_table_export(data: dict[str, Any], table_id: str, translate: Callable[[str], str]):
    """Return (headers, rows) for a table, using translated column labels."""
    cfg = _find_cfg(table_id, data)
    if not cfg:
        return [], []
    headers = [translate("col.index")] + [translate(c) for c in cfg["cols"]]
    n_cols = len(cfg["cols"])
    items = data.get(table_id) or []
    rows = []
    for idx, item in enumerate(items):
        cells = cfg["row"](item)
        # M11: pad / truncate to `n_cols` so a future contributor who adds
        # a custom row extractor can't silently misalign columns between
        # headers and rows on a CSV / Excel paste.
        cells = (cells + [""] * n_cols)[:n_cols]
        rows.append([str(idx + 1)] + ["" if v is None else str(v) for v in cells])
    return headers, rows


# ---------------------------------------------------------------------------
# Column type metadata
# ---------------------------------------------------------------------------
# Each table id maps to a per-column type tag so the GUI's "Apply Edits"
# path can parse the strings the user types back into the right Python
# type ("formations" → list[str], "from_bed_idx" → int, ...).
#
# Keys here are the same as ``data_keys`` on the table cfg (i.e. the
# actual dict key inside the result, not the i18n label).
#
# Types:
#   "str"     — leave as string (default; omit from the map)
#   "list"    — split by ";" and strip
#   "int"     — coerce to int, empty → None
#   "float"   — coerce to float, empty → None
COL_TYPES: dict[str, dict[str, str]] = {
    "sections": {
        "formations": "list",
    },
    "species_ranges": {},
    "biozones": {},
    "fossil_legend": {},
    "lithology_legend": {},
    "cross_beds": {
        "from_bed_idx": "int",
        "to_bed_idx": "int",
    },
    "other_fossils": {},
    # Abundance-diagram tables: all free-text (depth/abundance kept as
    # strings so units like "35%" or categorical "common" survive edits).
    "sites": {},
    "abundances": {},
    "zones": {},
}


def _coerce_cell(value: str, data_key: str, table_id: str) -> Any:
    """Convert a string from a table cell into the right Python type.

    The default ("str") leaves the value unchanged. ``list`` splits on
    ``;`` and trims; ``int``/``float`` parses and returns None on empty
    or invalid input.
    """
    type_map = COL_TYPES.get(table_id, {})
    t = type_map.get(data_key, "str")
    if t == "list":
        return [s.strip() for s in (value or "").split(";") if s.strip()]
    if t == "int":
        s = (value or "").strip()
        if not s:
            return None
        try:
            return int(s)
        except (TypeError, ValueError):
            return None
    if t == "float":
        s = (value or "").strip()
        if not s:
            return None
        try:
            return float(s)
        except (TypeError, ValueError):
            return None
    return value or ""


def apply_table_edits(
    data: dict[str, Any],
    table_id: str,
    rows: list[list[str]],
) -> dict[str, Any]:
    """Mutate ``data[table_id]`` from a list of cell rows.

    ``rows`` is the table widget's current state, one entry per row
    (the first cell is the auto-index column and is ignored). The shape
    of the produced items is inferred from the cfg: dict for everything
    except ``other_fossils`` (plain strings).
    """
    if not isinstance(data, dict):
        return data
    cfg = _find_cfg(table_id, data)
    if not cfg:
        return data
    data_keys = cfg.get("data_keys") or cfg["cols"]
    out: list[Any] = []
    for row in rows:
        # Skip empty placeholder rows (a Qt quirk: rowCount is 1 even
        # when the model is empty, so the last row is a phantom).
        if not row or all((c is None or str(c).strip() == "") for c in row[1:]):
            continue
        if table_id == "other_fossils":
            # Plain string list.
            txt = (row[1] if len(row) > 1 else "") or ""
            txt = str(txt).strip()
            if txt:
                out.append(txt)
            continue
        d: dict[str, Any] = {}
        for ci, dk in enumerate(data_keys, start=1):
            if dk == "agreement":
                # agreement is computed at merge time — ignore on edit.
                continue
            v = row[ci] if ci < len(row) else ""
            d[dk] = _coerce_cell(str(v or ""), dk, table_id)
        out.append(d)
    data[table_id] = out
    return data


def _sanitize_formula_cell(v: Any) -> Any:
    """Mitigate CSV/TSV formula injection (OWASP): prefix a cell whose first
    character is a formula trigger (= + - @) or a tab/CR/LF with a single
    quote so Excel/LibreOffice treats it as text, not an executable formula.
    An LLM-extracted (or attacker-crafted) value like =CMD(...) or
    =HYPERLINK(...) would otherwise execute on open.
    """
    if v is None:
        return v
    s = str(v)
    if s and s[0] in ("=", "+", "-", "@", "\t", "\r", "\n"):
        return "'" + s
    return v


def to_csv(headers: list[str], rows: list[list[str]]) -> str:
    """CSV text with a UTF-8 BOM so Excel reads CJK/Cyrillic correctly."""
    buf = io.StringIO()
    writer = csv.writer(buf, lineterminator="\r\n")
    writer.writerow([_sanitize_formula_cell(h) for h in headers])
    writer.writerows([[_sanitize_formula_cell(c) for c in row] for row in rows])
    return "﻿" + buf.getvalue()


def to_tsv(headers: list[str], rows: list[list[str]]) -> str:
    """TSV text; tabs/newlines inside cells collapsed to spaces."""
    def clean(v: Any) -> str:
        s = "" if v is None else str(v)
        s = " ".join(s.split())
        # Formula-injection mitigation (tab/CR/LF already collapsed above,
        # so only the = + - @ triggers remain possible).
        if s and s[0] in ("=", "+", "-", "@"):
            s = "'" + s
        return s

    lines = ["\t".join(clean(h) for h in headers)]
    for row in rows:
        lines.append("\t".join(clean(c) for c in row))
    return "\n".join(lines)


def result_to_json(data: dict[str, Any], source_file: str | None, timestamp: str | None) -> str:
    payload = {
        "extracted_at": timestamp,
        "source_file": source_file,
        "result": data,
    }
    import json

    return json.dumps(payload, ensure_ascii=False, indent=2)


# ---------------------------------------------------------------------------
# XLSX export
# ---------------------------------------------------------------------------

def _xlsx_sheet_name(title: str) -> str:
    """OpenPyXL sheet titles are limited to 31 characters (Unicode code points)
    and must not contain any of ``\\ / * ? [ ] :``. Invalid characters are
    replaced with underscores.

    The limit is per code point, not per UTF-8 byte. CJK characters (each
    1 code point, 3 bytes in UTF-8) are safe — Python 3 str indexing is by
    Unicode code point, so ``s[:31]`` never splits inside a character.
    We still validate by encoding to UTF-8 to confirm the result is <= 31 cp.
    """
    cleaned = "".join("_" if c in "\\/*?[]:" else c for c in (title or ""))
    cleaned = cleaned or "Sheet"
    # Python 3 str slicing is by Unicode code point, not byte. Enforce the
    # 31-code-point OpenPyXL limit safely for all scripts.
    if len(cleaned) > 31:
        cleaned = cleaned[:31]
    # Guard: if we somehow exceeded 31 code points after re-encoding
    # (should never happen with valid Unicode), truncate by code point count.
    if len(cleaned.encode("utf-8")) > 93:  # 31 * 3 bytes max per CJK
        chars = []
        byte_count = 0
        for c in cleaned:
            char_bytes = len(c.encode("utf-8"))
            if byte_count + char_bytes > 93:
                break
            chars.append(c)
            byte_count += char_bytes
        cleaned = "".join(chars)
    return cleaned


def to_xlsx(
    data: dict[str, Any],
    file_or_path: str | None = None,
    *,
    include_index: bool = True,
    translate: Callable[[str], str] | None = None,
) -> bytes | None:
    """XLSX with one sheet per table.

    Species name is italicized. Numeric-looking strings (e.g. ``"3m"``)
    are left as text — auto-converting them to numbers loses units.

    Returns
    -------
    bytes when ``file_or_path`` is None (caller writes to disk).
    None   when ``file_or_path`` is given and the file is written.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError as exc:
        raise RuntimeError(
            "XLSX export requires openpyxl. Install with: pip install openpyxl"
        ) from exc

    wb = Workbook()
    wb.remove(wb.active)

    def _t(key: str) -> str:
        # Translate i18n keys (e.g. "sec.speciesRanges", "col.species") so the
        # sheet names and headers match the user's language instead of showing
        # the raw keys. Falls back to the raw key when no translator is given.
        return translate(key) if translate else key

    bold = Font(bold=True)
    italic = Font(italic=True)
    header_fill = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
    thin = Side(border_style="thin", color="CBD5E1")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    italic_col_idx: dict[str, int] = {
        "species_ranges": 2,   # 1=# index column, 2=species
        "abundances": 2,       # 1=# index column, 2=taxon (genus → italic)
    }

    for cfg in get_configs_for_result(data):
        title_key = cfg.get("title_key") or cfg["id"]
        sheet_name = _xlsx_sheet_name(_t(title_key))
        ws = wb.create_sheet(title=sheet_name)

        headers = (["#"] if include_index else []) + [_t(c) for c in cfg["cols"]]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = bold
            cell.fill = header_fill
            cell.border = cell_border
            cell.alignment = center
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"

        items = (data or {}).get(cfg["id"]) or []
        italic_col = italic_col_idx.get(cfg["id"], -1)
        for idx, item in enumerate(items, start=1):
            if not isinstance(item, dict):
                continue
            row_values: list[Any] = []
            if include_index:
                row_values.append(idx)
            row_values.extend(cfg["row"](item))
            for ci, v in enumerate(row_values):
                row_values[ci] = "" if v is None else v
            ws.append(row_values)
            excel_row = idx + 1
            for ci, _ in enumerate(row_values, start=1):
                cell = ws.cell(row=excel_row, column=ci)
                cell.border = cell_border
                if ci == italic_col:
                    cell.font = italic
                cell.alignment = Alignment(
                    horizontal="left", vertical="top", wrap_text=True,
                )

        # Auto-size columns: min 10, max 60 (capped so a single long cell
        # doesn't blow the column out across multiple screen widths).
        for col_idx, col_cells in enumerate(ws.columns, start=1):
            try:
                length = max(
                    (len(str(c.value)) for c in col_cells if c.value is not None),
                    default=10,
                )
            except Exception:
                length = 10
            letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[letter].width = min(60, max(10, length + 2))

    if file_or_path:
        wb.save(file_or_path)
        return None
    from io import BytesIO
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
