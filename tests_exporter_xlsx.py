"""Tests for the XLSX exporter (rca_core.exporter.to_xlsx)."""

from __future__ import annotations

import io
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

_pass = 0
_fail = 0


def check(name, cond):
    global _pass, _fail
    if cond:
        _pass += 1
        print("PASS", name)
    else:
        _fail += 1
        print("FAIL", name)


def test_to_xlsx_bytes():
    try:
        from openpyxl import load_workbook
        from rca_core.exporter import to_xlsx
    except ImportError as exc:
        check("xlsx-import", False)
        print("SKIP: openpyxl not available:", exc)
        return
    check("xlsx-import", True)

    data = {
        "sections": [
            {"name": "Sec A", "age_range": "Permian", "formations": ["F1", "F2"],
             "formation_thickness_m": "10m", "coordinates": "31N"},
        ],
        "species_ranges": [
            {"species": "Sp. x", "section": "Sec A", "range_base": "Bed 1",
             "range_top": "Bed 3", "biozone": "Z1"},
            {"species": "Sp. y", "section": "Sec A", "range_base": "Bed 2",
             "range_top": "Bed 4", "biozone": ""},
        ],
        "biozones": [{"name": "Z1", "age": "Late", "thickness_m": "3m"}],
        "other_fossils": ["Ammonoid: X"],
        "confidence": 0.9,
    }
    blob = to_xlsx(data)
    check("xlsx-bytes-not-empty", blob and len(blob) > 0)
    check("xlsx-zip-magic", blob[:2] == b"PK")

    # Round-trip
    wb = load_workbook(io.BytesIO(blob))
    check("xlsx-sheet-count", len(wb.sheetnames) == 4)
    # The species name should be italic.
    sp_sheet = wb["Species Ranges (Species Ranges)"] if "Species Ranges (Species Ranges)" in wb.sheetnames else wb[wb.sheetnames[1]]
    italic = sp_sheet.cell(row=2, column=2).font.italic
    check("xlsx-species-italic", italic is True)


def test_to_xlsx_to_file(tmp_path=None):
    from rca_core.exporter import to_xlsx
    import tempfile
    with tempfile.TemporaryDirectory() as td:
        data = {
            "sections": [{"name": "S1", "age_range": "P", "formations": [],
                          "formation_thickness_m": "", "coordinates": ""}],
            "species_ranges": [{"species": "S. x", "section": "S1", "range_base": "",
                                 "range_top": "", "biozone": ""}],
            "biozones": [],
            "other_fossils": [],
        }
        path = os.path.join(td, "out.xlsx")
        ret = to_xlsx(data, file_or_path=path)
        check("xlsx-to-file-returns-None", ret is None)
        check("xlsx-to-file-exists", os.path.isfile(path))
        check("xlsx-to-file-size", os.path.getsize(path) > 0)


def test_to_xlsx_sheet_name_sanitization():
    """Sheet name with illegal chars must be sanitized to underscores."""
    from rca_core.exporter import _xlsx_sheet_name
    check("xlsx-name-strip", _xlsx_sheet_name("a/b\\c?d*e[f]g:h") == "a_b_c_d_e_f_g_h")
    check("xlsx-name-truncate", len(_xlsx_sheet_name("x" * 100)) == 31)
    # CJK boundary: truncating by code point count must not cut a
    # multi-byte character in half (the previous byte-based slice
    # could land mid-character for a Chinese title). Verify by
    # round-tripping through len() and absence of lone surrogates.
    cjk = "化石记录 (Abundances) — 延限图"
    out = _xlsx_sheet_name(cjk)
    check("xlsx-name-cjk-len", len(out) <= 31)
    # No unpaired surrogate halves (Unicode "lone surrogates" range).
    lone_surr = any(0xD800 <= ord(c) <= 0xDFFF for c in out)
    check("xlsx-name-cjk-no-lone-surrogate", not lone_surr)


def test_body_too_large_key_present():
    """The 413 error key must be present in every supported i18n language so
    the frontend t() lookup never falls back to the raw key string."""
    from rca_core.i18n import Translator, TRANSLATIONS
    for lang in ("zh", "en", "ja"):
        # Translator.t returns the key itself when missing; check that
        # we instead get a real (non-key) translated string.
        val = Translator(lang).t("err.bodyTooLarge")
        check(f"err-bodyTooLarge-{lang}-translated",
              val and val != "err.bodyTooLarge")


def test_missing_dependency_error():
    """If openpyxl is missing, to_xlsx must raise a clear RuntimeError, not ImportError."""
    from rca_core.exporter import to_xlsx
    import builtins
    orig = builtins.__import__
    def fake_import(name, *a, **kw):
        if name == "openpyxl" or name.startswith("openpyxl."):
            raise ImportError("simulated missing openpyxl")
        return orig(name, *a, **kw)
    builtins.__import__ = fake_import
    try:
        try:
            to_xlsx({"sections": [], "species_ranges": [], "biozones": [], "other_fossils": []})
            check("xlsx-missing-raises", False)
        except RuntimeError as exc:
            check("xlsx-missing-raises", "openpyxl" in str(exc))
    finally:
        builtins.__import__ = orig


test_to_xlsx_bytes()
test_to_xlsx_to_file()
test_to_xlsx_sheet_name_sanitization()
test_body_too_large_key_present()
test_missing_dependency_error()
print("--- %d passed, %d failed ---" % (_pass, _fail))
sys.exit(1 if _fail else 0)

